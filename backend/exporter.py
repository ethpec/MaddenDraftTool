"""xlsx exporters for a finished (or in-progress) draft.

Three workbooks are produced:

1. **DraftPickOutcome.xlsx** — one row per pick with the player chosen,
   the team that ended up making the selection, and round/pick numbers.
2. **DraftPicks_updated.xlsx** — same shape as the input DraftPicks.xlsx
   but with ``CurrentTeam`` rewritten for any traded picks and
   ``SelectedPlayer`` populated where applicable. This is the file the
   game would re-import.
3. **Trades.xlsx** — chronological log of every trade for testing.

Outputs land under ``Files/<year>/Exports/draft_<timestamp>/``.
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any
import zipfile

import openpyxl

from .data_loader import resolve_year_folder
from .draft_state import DraftSession, _pick_to_dict
from .logic import pick_value


def export_session(session: DraftSession, year: str | int | None) -> dict[str, str]:
    """Write all three xlsx outputs and return their paths."""
    folder = resolve_year_folder(year)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_dir = folder / "Exports" / f"draft_{stamp}"
    out_dir.mkdir(parents=True, exist_ok=True)

    outcome_path = out_dir / "DraftPickOutcome.xlsx"
    picks_path = out_dir / "DraftPicks_updated.xlsx"
    trades_path = out_dir / "Trades.xlsx"

    _write_outcome(session, outcome_path)
    _write_updated_picks(session, year, picks_path)
    _write_trades(session, trades_path)

    return {
        "outcome": str(outcome_path),
        "picks": str(picks_path),
        "trades": str(trades_path),
        "folder": str(out_dir),
    }


def export_session_zip(session: DraftSession, year: str | int | None) -> dict[str, str]:
    """Write one export folder and zip the three workbook outputs."""
    paths = export_session(session, year)
    out_dir = Path(paths["folder"])
    zip_path = out_dir / "DraftExport.zip"
    with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        for kind in ("outcome", "picks", "trades"):
            source = Path(paths[kind])
            zf.write(source, arcname=source.name)
    paths["zip"] = str(zip_path)
    return paths


def _write_outcome(session: DraftSession, path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "DraftPickOutcome"
    ws.append([
        "Overall", "Round", "PickInRound", "OriginalTeam", "DraftingTeam",
        "PlayerID", "PlayerName",
    ])
    for pick in session.board():
        ws.append([
            pick.overall, pick.round_1, pick.pick_in_round_1,
            pick.original_team, pick.current_team,
            pick.selected_player_id, pick.selected_player_name,
        ])
    wb.save(path)


def _write_updated_picks(session: DraftSession, year: str | int | None, path: Path) -> None:
    """Rewrite the original DraftPicks.xlsx with updated CurrentTeam + selections.

    We re-open the source file to preserve column order/format, mutate
    CurrentTeam (and SelectedPlayer for current-year picks), and save under
    a new name. Both current-year (YearOffset=0) and future picks
    (YearOffset=1) are updated so traded next-year picks reflect their new
    owner on re-import.
    """
    source = resolve_year_folder(year) / "DraftPicks.xlsx"
    wb = openpyxl.load_workbook(source)
    ws = wb.active

    # Index picks by (year_offset, Round, PickNumber) — both current and future.
    by_key: dict[tuple[int, int, int], Any] = {}
    for pick in session.board():
        by_key[(0, pick.round_1 - 1, pick.pick_in_round_1 - 1)] = pick
    for pick in session.future_picks():
        by_key[(1, pick.round_1 - 1, pick.pick_in_round_1 - 1)] = pick

    # Find header positions.
    headers = [c.value for c in ws[1]]
    h = {name: headers.index(name) + 1 for name in headers if name}

    # team_info is {TeamNumber: TeamName}; reverse it for the encode lookup.
    # GMInfo["TeamIndex"] is a different numbering and must not be used here.
    name_to_idx = {
        name: int(num)
        for num, name in session.data.get("team_info", {}).items()
    }

    for row_i in range(2, ws.max_row + 1):
        year_off = int(ws.cell(row_i, h["YearOffset"]).value or 0)
        rnd = ws.cell(row_i, h["Round"]).value or 0
        pk = ws.cell(row_i, h["PickNumber"]).value or 0
        pick = by_key.get((year_off, rnd, pk))
        if not pick:
            continue
        idx = name_to_idx.get(pick.current_team)
        if idx is not None:
            original = ws.cell(row_i, h["CurrentTeam"]).value
            ws.cell(row_i, h["CurrentTeam"]).value = _encode_team_id(idx, original)
        if pick.selected_player_id:
            ws.cell(row_i, h["SelectedPlayer"]).value = pick.selected_player_id
    wb.save(path)


def _write_trades(session: DraftSession, path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Trades"
    ws.append([
        "TradeID", "InitiatedBy", "HeadlinePickOverall",
        "TeamA", "TeamB", "TeamASends", "TeamBSends",
        "TeamAValueSent", "TeamBValueSent",
    ])
    pv_table = session.data.get("pick_values", {})
    for t in session.trade_log():
        ws.append([
            t.trade_id, t.initiated_by, t.overall_pick_traded,
            t.team_a, t.team_b,
            _summarize_picks(t.team_a_sends),
            _summarize_picks(t.team_b_sends),
            _total_pick_value(t.team_a_sends, pv_table),
            _total_pick_value(t.team_b_sends, pv_table),
        ])
    wb.save(path)


def _summarize_picks(picks: list[dict[str, Any]]) -> str:
    return "; ".join(f"R{p['round']}.{p['pick_in_round']} (overall {p['overall']})" for p in picks)


def _total_pick_value(picks: list[dict[str, Any]], pv_table: dict[str, Any]) -> float:
    return round(sum(pick_value(p, pv_table) for p in picks), 1)


def _encode_team_id(team_index: int, original_value: Any = None) -> str:
    """Reconstruct the 32-bit binary team ID used by DraftPicks.xlsx.

    The high 24 bits are a version-specific prefix that varies across Madden
    releases. We read it from the row's existing value so the output always
    matches what the game expects, regardless of version.
    """
    s = str(original_value) if original_value is not None else ""
    prefix = s[:24] if len(s) == 32 and all(c in "01" for c in s) else "001011010100101000000000"
    return prefix + format(team_index & 0xFF, "08b")
