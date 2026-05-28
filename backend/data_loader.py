"""Data loader for Madden Draft Tool.

Resolves a year selection to a folder under ``Files/`` and parses the
xlsx/json source files into in-memory Python dicts the rest of the
backend can consume.

Folder convention: ``Files/<year>/``. If a folder for the requested year
does not exist, falls back to ``Files/TestFiles/`` so the tool stays
usable while only test data is present.
"""

from __future__ import annotations

import copy
import json
import re
import threading
from pathlib import Path
from typing import Any

import openpyxl


REPO_ROOT = Path(__file__).resolve().parent.parent
FILES_ROOT = REPO_ROOT / "Files"
FALLBACK_FOLDER_NAME = "TestFiles"
NFL_LOGOS_DIR = REPO_ROOT / "static" / "nfl_logos"
COLLEGE_LOGOS_DIR = REPO_ROOT / "static" / "college_logos"
PORTRAITS_DIR = REPO_ROOT / "static" / "portraits"
DATA_CACHE_DIRNAME = ".cache"
PLAYER_CACHE_VERSION = 4
LOAD_ALL_CACHE_VERSION = 6
_LOAD_ALL_CACHE_LOCK = threading.Lock()
_LOAD_ALL_CACHE: dict[tuple[Any, ...], dict[str, Any]] = {}


# Columns from Player.xlsx we actually care about for needs / roster OVR.
# Loading all 271 columns is wasteful; this list is the minimal join we
# need for now, expand as logic functions need more.
PLAYER_COLUMNS_OF_INTEREST: tuple[str, ...] = (
    "FirstName",
    "LastName",
    "Position",
    "OverallRating",
    "Age",
    "TeamIndex",
    "ContractStatus",
    "PLYR_DRAFTTEAM",
    "YearDrafted",
    "College",
    "PLYR_ASSETNAME",
    "PLYR_DRAFTROUND",
    "PLYR_DRAFTPICK",
    "PLYR_PORTRAIT",
    "TraitDevelopment",
)


def slugify(name: str) -> str:
    """Convert a team or college name to the on-disk logo slug.

    Logos in ``static/nfl_logos`` and ``static/college_logos`` use
    lowercase-hyphenated filenames (``pittsburgh-steelers.png``,
    ``alabama-a-and-m.png``). We mirror that convention: lowercase,
    ampersands become ``and``, drop apostrophes, anything else
    non-alphanumeric -> hyphen.
    """
    if not name:
        return ""
    s = str(name).lower().replace("&", " and ").replace("'", "")
    s = re.sub(r"[^a-z0-9]+", "-", s)
    return s.strip("-")


def build_nfl_logo_map() -> dict[str, str]:
    """Map team nickname (matches GMInfo TeamName) -> logo URL path.

    NFL logos are named ``city-nickname.png`` (e.g.
    ``pittsburgh-steelers.png``). GMInfo only carries the nickname, so
    we look for any file whose slug ends with the nickname slug.
    Returns ``{team_name: "/static/nfl_logos/foo.png"}`` and silently
    omits teams whose logo isn't on disk.
    """
    if not NFL_LOGOS_DIR.is_dir():
        return {}
    files = [f.name for f in NFL_LOGOS_DIR.iterdir() if f.is_file() and f.suffix == ".png"]
    out: dict[str, str] = {}
    # Hardcoded full-name slugs for the 32 teams; this is the most reliable
    # mapping (some nicknames like "Cardinals" or "Panthers" exist in
    # multiple cities historically, but GMInfo always means the current
    # franchise so we fix the city explicitly).
    name_to_slug = {
        "Bills": "buffalo-bills", "Dolphins": "miami-dolphins",
        "Patriots": "new-england-patriots", "Jets": "new-york-jets",
        "Ravens": "baltimore-ravens", "Bengals": "cincinnati-bengals",
        "Browns": "cleveland-browns", "Steelers": "pittsburgh-steelers",
        "Texans": "houston-texans", "Colts": "indianapolis-colts",
        "Jaguars": "jacksonville-jaguars", "Titans": "tennessee-titans",
        "Broncos": "denver-broncos", "Chiefs": "kansas-city-chiefs",
        "Chargers": "los-angeles-chargers", "Raiders": "las-vegas-raiders",
        "Cowboys": "dallas-cowboys", "Giants": "new-york-giants",
        "Eagles": "philadelphia-eagles", "Commanders": "washington-commanders",
        "Bears": "chicago-bears", "Lions": "detroit-lions",
        "Packers": "green-bay-packers", "Vikings": "minnesota-vikings",
        "Falcons": "atlanta-falcons", "Panthers": "carolina-panthers",
        "Saints": "new-orleans-saints", "Buccaneers": "tampa-bay-buccaneers",
        "Cardinals": "arizona-cardinals", "Rams": "los-angeles-rams",
        "Seahawks": "seattle-seahawks", "49ers": "san-francisco-49ers",
    }
    file_set = set(files)
    for team, slug in name_to_slug.items():
        fname = slug + ".png"
        if fname in file_set:
            out[team] = "/static/nfl_logos/" + fname
    return out


def build_portrait_files() -> list[str]:
    """Return a sorted list of portrait filenames from ``static/portraits``.

    There's no published Madden mapping of PLYR_PORTRAIT IDs to these
    files, so the API layer hashes the player's portrait ID (or name)
    into this list to pick a stable per-player portrait. Same player ->
    same portrait every render.
    """
    if not PORTRAITS_DIR.is_dir():
        return []
    return sorted(f.name for f in PORTRAITS_DIR.iterdir()
                  if f.is_file() and f.suffix.lower() == ".png")


def build_college_logo_map(colleges: Any) -> dict[str, str]:
    """Map collegeName -> logo URL path; only includes names whose file exists.

    ``colleges`` is the parsed all_colleges.json (list of
    ``{collegeName, binaryReference}``). Roughly 380/488 colleges have
    matching files in ``static/college_logos``; the rest get no logo
    and the UI will show a placeholder.
    """
    if not COLLEGE_LOGOS_DIR.is_dir() or not colleges:
        return {}
    files = {f.name for f in COLLEGE_LOGOS_DIR.iterdir()
             if f.is_file() and f.suffix == ".png"}
    out: dict[str, str] = {}
    for c in colleges:
        name = c.get("collegeName")
        if not name:
            continue
        fname = slugify(name) + ".png"
        if fname in files:
            out[name] = "/static/college_logos/" + fname
    return out


def build_college_id_to_name(colleges: Any) -> dict[str, str]:
    """Decode Player.xlsx ``College`` column (binary string) -> college name."""
    if not colleges:
        return {}
    return {c["binaryReference"]: c["collegeName"]
            for c in colleges if c.get("binaryReference")}


def resolve_year_folder(year: str | int | None) -> Path:
    """Return the folder path for a given draft year.

    Looks for ``Files/<year>/``. If that does not exist, returns
    ``Files/TestFiles/`` so the rest of the loader still works against
    sample data.
    """
    if year is not None:
        candidate = FILES_ROOT / str(year)
        if candidate.is_dir():
            return candidate
    return FILES_ROOT / FALLBACK_FOLDER_NAME


def list_available_years() -> list[dict[str, Any]]:
    """Return entries describing every draft folder under ``Files/``.

    Each entry has ``{"year": str, "is_test": bool, "path": str}``.
    """
    if not FILES_ROOT.is_dir():
        return []
    entries: list[dict[str, Any]] = []
    for child in sorted(FILES_ROOT.iterdir()):
        if not child.is_dir():
            continue
        if child.name.lower() == "exports":
            continue
        entries.append({
            "year": child.name,
            "is_test": child.name == FALLBACK_FOLDER_NAME,
            "path": str(child),
        })
    return entries


def _read_sheet(path: Path, sheet_name: str | None = None) -> tuple[list[str], list[list[Any]]]:
    """Open an xlsx file and return ``(headers, rows)`` for the chosen sheet.

    If ``sheet_name`` is None, the active/first sheet is used.
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active
    iterator = ws.iter_rows(values_only=True)
    try:
        headers_tuple = next(iterator)
    except StopIteration:
        wb.close()
        return [], []
    headers = [str(h) if h is not None else "" for h in headers_tuple]
    rows: list[list[Any]] = []
    for row in iterator:
        if row is None or all(c is None for c in row):
            continue
        rows.append(list(row))
    wb.close()
    return headers, rows


def _rows_to_dicts(headers: list[str], rows: list[list[Any]]) -> list[dict[str, Any]]:
    return [
        {headers[i]: (row[i] if i < len(row) else None) for i in range(len(headers))}
        for row in rows
    ]


def _file_signature(path: Path) -> dict[str, Any]:
    stat = path.stat()
    return {
        "path": str(path),
        "mtime_ns": stat.st_mtime_ns,
        "size": stat.st_size,
    }


def _folder_signature(folder: Path) -> tuple[Any, ...]:
    """Return a cheap invalidation key for source data + logo assets."""
    source_names = (
        "BigBoard.xlsx",
        "Player.xlsx",
        "DraftPicks.xlsx",
        "GMInfo.xlsx",
        "PositionNeeds.xlsx",
        "DraftPickValue.xlsx",
        "DraftMaxPerPositionGroup.xlsx",
        "Draft_LetterGrades.xlsx",
        "all_colleges.json",
    )
    parts: list[Any] = [LOAD_ALL_CACHE_VERSION]
    for name in source_names:
        path = folder / name
        parts.append((name, path.stat().st_mtime_ns, path.stat().st_size) if path.is_file() else (name, None, None))
    for logo_dir in (NFL_LOGOS_DIR, COLLEGE_LOGOS_DIR):
        logo_parts = []
        if logo_dir.is_dir():
            for path in sorted(logo_dir.glob("*.png")):
                stat = path.stat()
                logo_parts.append((path.name, stat.st_mtime_ns, stat.st_size))
        parts.append((str(logo_dir), tuple(logo_parts)))
    return tuple(parts)


def load_big_board(folder: Path) -> dict[str, Any]:
    """Load BigBoard.xlsx -> draftable rookies.

    Columns used: FirstName, LastName, Player_ID, Drafted, BigBoardRank,
    PLYR_DRAFTROUND, PLYR_DRAFTPICK, ProspectType. Per-team ranking
    columns are not needed — team boards are generated via noise in
    compute_team_big_board using BigBoardRank as the shared baseline.
    """
    headers, rows = _read_sheet(folder / "BigBoard.xlsx")
    keep_cols = {"FirstName", "LastName", "Player_ID", "Drafted",
                 "BigBoardRank", "PLYR_DRAFTROUND", "PLYR_DRAFTPICK", "ProspectType",
                 "GenericHeadAssetName"}
    players: list[dict[str, Any]] = []
    for row in rows:
        record = {headers[i]: row[i] for i in range(len(headers))
                  if headers[i] in keep_cols}
        record["drafted"] = record.get("Drafted") not in (None, "", 0, False)
        players.append(record)
    return {"players": players}


def load_players(folder: Path) -> list[dict[str, Any]]:
    """Load Player.xlsx but project to columns of interest only.

    Player.xlsx is ~32MB with 271 columns; we never need most of them in
    the draft tool. We keep just the columns useful for computing team
    needs and current roster OVR.
    """
    path = folder / "Player.xlsx"
    cache_path = folder / DATA_CACHE_DIRNAME / "Player.projected.json"
    metadata = {
        "version": PLAYER_CACHE_VERSION,
        "source": _file_signature(path),
        "columns": list(PLAYER_COLUMNS_OF_INTEREST),
    }
    if cache_path.is_file():
        try:
            with cache_path.open() as f:
                cached = json.load(f)
            if cached.get("metadata") == metadata:
                return cached.get("players", [])
        except (OSError, json.JSONDecodeError):
            pass

    headers, rows = _read_sheet(path)
    keep_idx = {h: headers.index(h) for h in PLAYER_COLUMNS_OF_INTEREST if h in headers}
    out: list[dict[str, Any]] = []
    for i, row in enumerate(rows):
        rec = {col: row[idx] for col, idx in keep_idx.items() if idx < len(row)}
        rec["player_table_row"] = i + 2  # Excel row number (header=1, first data row=2)
        out.append(rec)
    try:
        cache_path.parent.mkdir(parents=True, exist_ok=True)
        tmp_path = cache_path.with_suffix(cache_path.suffix + ".tmp")
        with tmp_path.open("w") as f:
            json.dump({"metadata": metadata, "players": out}, f)
        tmp_path.replace(cache_path)
    except OSError:
        pass
    return out


# Combine/Pro Day drill columns we extract, in display order.
_COMBINE_DRILLS: list[str] = [
    "FortyYardDash", "TwentyYardShuttle", "ThreeConeDrill",
    "VerticalJump", "BroadJump", "BenchPress",
]

# True = lower value is better (times); False = higher is better (jumps/reps).
_DRILL_ASCENDING: dict[str, bool] = {
    "FortyYardDash": True, "TwentyYardShuttle": True, "ThreeConeDrill": True,
    "VerticalJump": False, "BroadJump": False, "BenchPress": False,
}

# Group split Madden positions together for combine ranking so "LT" and "RT"
# both compete in the same "OT" pool, etc.
_RANK_POSITION_GROUP: dict[str, str] = {
    "LT": "OT", "RT": "OT",
    "LG": "OG", "RG": "OG",
    "LE": "EDGE", "RE": "EDGE",
    "LOLB": "OLB", "ROLB": "OLB",
    "MLB": "ILB",
    "HB": "RB",
}


def _compute_combine_ranks(bb_players: list[dict[str, Any]]) -> None:
    """Add rank dicts to each player's combine_data in-place.

    Groups players by position (split positions merged via _RANK_POSITION_GROUP),
    then for each drill ranks using standard competition ranking (1224): ties
    share a rank, next rank skips. Stores results as:
      combine_data["combine_ranks"][drill] = {"rank": N, "total": M}
      combine_data["pro_day_ranks"][drill] = {"rank": N, "total": M}
    """
    # Build position groups
    groups: dict[str, list[dict[str, Any]]] = {}
    for p in bb_players:
        pos = _RANK_POSITION_GROUP.get(p.get("position") or "", p.get("position") or "UNK")
        groups.setdefault(pos, []).append(p)

    for players in groups.values():
        for src, ranks_key in (("combine", "combine_ranks"), ("pro_day", "pro_day_ranks")):
            for drill, ascending in _DRILL_ASCENDING.items():
                valid: list[tuple[Any, dict[str, Any]]] = []
                for p in players:
                    cd = p.get("combine_data")
                    if not cd:
                        continue
                    val = cd[src].get(drill)
                    if val is not None:
                        valid.append((val, p))
                if not valid:
                    continue
                total = len(valid)
                valid.sort(key=lambda x: x[0], reverse=not ascending)
                # Standard competition ranking (1224)
                i = 0
                current_rank = 1
                while i < len(valid):
                    j = i
                    while j < len(valid) and valid[j][0] == valid[i][0]:
                        j += 1
                    for k in range(i, j):
                        p = valid[k][1]
                        cd = p["combine_data"]
                        cd.setdefault(ranks_key, {})[drill] = {
                            "rank": current_rank, "total": total,
                        }
                    current_rank += j - i
                    i = j


def load_combine_pro_day(folder: Path) -> dict[int, dict[str, Any]]:
    """Load CombineProDayData.xlsx -> {PlayerTableRow: {combine: {...}, pro_day: {...}}}.

    Keyed by PlayerTableRow (int). Each inner dict has raw integer values for
    the six drills (or None if not recorded). Returns {} if the file is absent.
    """
    path = folder / "CombineProDayData.xlsx"
    if not path.is_file():
        return {}
    headers, rows = _read_sheet(path)
    dicts = _rows_to_dicts(headers, rows)
    out: dict[int, dict[str, Any]] = {}
    for row in dicts:
        ptr = row.get("PlayerTableRow")
        if ptr is None:
            continue
        ptr = int(ptr)
        combine = {drill: row.get(f"Combine{drill}") for drill in _COMBINE_DRILLS}
        pro_day = {drill: row.get(f"ProDay{drill}") for drill in _COMBINE_DRILLS}
        out[ptr] = {"combine": combine, "pro_day": pro_day}
    return out


def _decode_team_number(value: Any) -> int | None:
    """DraftPicks.xlsx stores team identity as a 32-bit binary string.

    The lowest 8 bits are the TeamNumber from TeamInfo.xlsx. We decode
    just that byte so callers can look the team name up in team_info.
    """
    if value is None:
        return None
    s = str(value)
    if len(s) != 32 or any(c not in "01" for c in s):
        return None
    return int(s[-8:], 2)


def load_team_info(folder: Path) -> dict[int, str]:
    """Load TeamInfo.xlsx -> {TeamNumber: TeamName} for real NFL teams only.

    Filters out non-team rows (AFC, NFC, Free Agents, Hall Of Fame, NFL
    Greats) whose TeamIndex >= 32.
    """
    headers, rows = _read_sheet(folder / "TeamInfo.xlsx")
    dicts = _rows_to_dicts(headers, rows)
    out: dict[int, str] = {}
    for d in dicts:
        team_num = d.get("TeamNumber")
        name = d.get("TeamName")
        team_idx = d.get("TeamIndex")
        if team_num is None or not name or team_idx is None:
            continue
        if int(team_idx) >= 32:
            continue
        out[int(team_num)] = str(name)
    return out


def load_draft_picks(folder: Path) -> list[dict[str, Any]]:
    """Load DraftPicks.xlsx.

    Columns: SelectedPlayer, OriginalTeam, CurrentTeam, YearOffset,
    Round, PickNumber. ``Round`` / ``PickNumber`` are 0-indexed in the
    source. Adds 1-indexed convenience fields and decoded TeamNumber ints
    (from TeamInfo.xlsx) for OriginalTeam / CurrentTeam.
    """
    headers, rows = _read_sheet(folder / "DraftPicks.xlsx")
    dicts = _rows_to_dicts(headers, rows)
    for d in dicts:
        d["round_1"] = (d.get("Round") or 0) + 1
        d["pick_1"] = (d.get("PickNumber") or 0) + 1
        d["OriginalTeamIndex"] = _decode_team_number(d.get("OriginalTeam"))
        d["CurrentTeamIndex"] = _decode_team_number(d.get("CurrentTeam"))
    return dicts


def load_gm_info(folder: Path) -> list[dict[str, Any]]:
    """Load GMInfo.xlsx -> per-team GM trait scores.

    Trailing legend rows (where TeamName is None) are dropped.
    """
    headers, rows = _read_sheet(folder / "GMInfo.xlsx")
    dicts = _rows_to_dicts(headers, rows)
    return [d for d in dicts if d.get("TeamName")]


def load_position_needs(folder: Path) -> list[dict[str, Any]]:
    """Load PositionNeeds.xlsx (TierNeedLogic sheet)."""
    headers, rows = _read_sheet(folder / "PositionNeeds.xlsx", "TierNeedLogic")
    return _rows_to_dicts(headers, rows)


def load_draft_pick_value(folder: Path) -> dict[str, Any]:
    """Load DraftPickValue.xlsx (Current + Future sheets).

    Also builds two O(1) lookups so hot paths don't scan the sheet on every
    call:

    * ``by_pick``: (slot_1indexed, year_offset) -> chart value. Matches
      ``logic.pick_value`` semantics — current-year picks read the Current
      sheet's ``Value`` column, future picks read the Future sheet's ``Value``.
    * ``next_year_by_slot``: slot_1indexed -> ValueNextYear from the Current
      sheet. Used by ``app._pick_point_value`` for displaying future picks'
      discounted value in the UI annotations.
    """
    cur_h, cur_r = _read_sheet(folder / "DraftPickValue.xlsx", "Current")
    fut_h, fut_r = _read_sheet(folder / "DraftPickValue.xlsx", "Future")
    current = _rows_to_dicts(cur_h, cur_r)
    future = _rows_to_dicts(fut_h, fut_r)
    by_pick: dict[tuple[int, int], float] = {}
    next_year_by_slot: dict[int, float] = {}
    for row in current:
        slot_0 = row.get("Pick")
        if slot_0 is None:
            continue
        slot_1 = int(slot_0) + 1
        if row.get("Value") is not None:
            by_pick[(slot_1, 0)] = float(row["Value"])
        if row.get("ValueNextYear") is not None:
            next_year_by_slot[slot_1] = float(row["ValueNextYear"])
    for row in future:
        slot_0 = row.get("Pick")
        if slot_0 is None:
            continue
        slot_1 = int(slot_0) + 1
        if row.get("Value") is not None:
            by_pick[(slot_1, 1)] = float(row["Value"])
    return {
        "current": current,
        "future": future,
        "by_pick": by_pick,
        "next_year_by_slot": next_year_by_slot,
    }



def load_max_per_position_group(folder: Path) -> dict[str, Any]:
    """Load DraftMaxPerPositionGroup.xlsx.

    Returns {
        "pos_to_group": {position: group_name},   # Position sheet
        "max_per_group": {group_name: max_int},    # PositionGroup sheet
    }
    Returns empty dicts for both if the file is absent.
    """
    path = folder / "DraftMaxPerPositionGroup.xlsx"
    if not path.is_file():
        return {"pos_to_group": {}, "max_per_group": {}}
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    pos_to_group: dict[str, str] = {}
    max_per_group: dict[str, int] = {}
    if "Position" in wb.sheetnames:
        ws = wb["Position"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            pos, grp = row[0], row[1]
            if pos and grp:
                pos_to_group[str(pos)] = str(grp)
    if "PositionGroup" in wb.sheetnames:
        ws = wb["PositionGroup"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            grp, mx = row[0], row[1]
            if grp and mx is not None:
                max_per_group[str(grp)] = int(mx)
    wb.close()
    return {"pos_to_group": pos_to_group, "max_per_group": max_per_group}


def load_letter_grades(folder: Path) -> dict[tuple[Any, ...], dict[str, Any]]:
    """Load Draft_LetterGrades.xlsx -> grade records keyed for safe joins.

    The workbook has an "All" overview sheet and per-position sheets (QB, RB,
    WR, TE, OL, DL, LB, CB, SAF, ST) whose columns mix three kinds of fields:
    bio (Rank/Round/Pick/FirstName/LastName/Height/Weight/Age/Position), a
    "Star" header tier (only present on the position sheets), and the rest are
    attributes — letter grades like A/B/F or star ratings like "6-Great".

    Each record is stored under TWO keys so callers can match strictly or fall
    back if positions disagree across data sources:
      ("fl", FirstName, LastName)               -- loose
      ("flp", FirstName, LastName, Position)    -- strict
    The "All" sheet is a duplicate of bio info we already have from BigBoard,
    so it's skipped.
    """
    path = folder / "Draft_LetterGrades.xlsx"
    if not path.is_file():
        return {}
    BIO_COLS = {"Rank", "Round", "Pick", "FirstName", "LastName",
                "Height", "Weight", "Age", "Position", "Star"}
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    out: dict[tuple[Any, ...], dict[str, Any]] = {}
    for sn in wb.sheetnames:
        if sn == "All":
            continue
        ws = wb[sn]
        it = ws.iter_rows(values_only=True)
        try:
            headers = list(next(it))
        except StopIteration:
            continue
        for row in it:
            if not row or row[headers.index("FirstName")] is None:
                continue
            first = row[headers.index("FirstName")]
            last = row[headers.index("LastName")]
            position = row[headers.index("Position")] if "Position" in headers else None
            rec: dict[str, Any] = {
                "position_group": sn,
                "position": position,
                "height": row[headers.index("Height")] if "Height" in headers else None,
                "weight": row[headers.index("Weight")] if "Weight" in headers else None,
                "age": row[headers.index("Age")] if "Age" in headers else None,
                "star": row[headers.index("Star")] if "Star" in headers else None,
                "attributes": {},
            }
            for i, h in enumerate(headers):
                if not h or h in BIO_COLS:
                    continue
                val = row[i] if i < len(row) else None
                if val is not None and val != "":
                    rec["attributes"][h] = val
            out[("flp", first, last, position)] = rec
            out.setdefault(("fl", first, last), rec)
    wb.close()
    return out


def load_colleges(folder: Path) -> Any:
    """Load all_colleges.json if present."""
    path = folder / "all_colleges.json"
    if not path.is_file():
        return None
    with path.open() as f:
        return json.load(f)


def load_all(year: str | int | None) -> dict[str, Any]:
    """Top-level entry point. Returns a single dict with every parsed source.

    Joins college names onto BigBoard rows by Player_ID -> Player.xlsx
    so the frontend can render college logos without doing the join
    itself.
    """
    folder = resolve_year_folder(year)
    cache_key = (str(year) if year is not None else None, str(folder), _folder_signature(folder))
    with _LOAD_ALL_CACHE_LOCK:
        cached = _LOAD_ALL_CACHE.get(cache_key)
    if cached is not None:
        return copy.deepcopy(cached)

    big_board = load_big_board(folder)
    players = load_players(folder)
    combine_data = load_combine_pro_day(folder)
    letter_grades = load_letter_grades(folder)
    colleges = load_colleges(folder)
    college_by_id = build_college_id_to_name(colleges)
    college_logo_map = build_college_logo_map(colleges)
    nfl_logo_map = build_nfl_logo_map()

    # Annotate big board players with their college name + logo URL by
    # looking them up in Player.xlsx by FirstName+LastName (BigBoard
    # uses the same identifier shape — "FirstName LastName_id" — but
    # Player.xlsx joins more reliably on the (First, Last) tuple).
    player_lookup: dict[tuple, dict[str, Any]] = {}
    for pl in players:
        key = (pl.get("FirstName"), pl.get("LastName"),
               pl.get("PLYR_DRAFTROUND"), pl.get("PLYR_DRAFTPICK"))
        if key not in player_lookup:
            player_lookup[key] = pl
    for p in big_board["players"]:
        key = (p.get("FirstName"), p.get("LastName"),
               p.get("PLYR_DRAFTROUND"), p.get("PLYR_DRAFTPICK"))
        match = player_lookup.get(key)
        if match:
            college_id = match.get("College")
            college_name = college_by_id.get(college_id)
            p["college"] = college_name
            p["position"] = match.get("Position")
            if college_name:
                p["college_logo"] = college_logo_map.get(college_name)
            # Propagate the Madden portrait ID; the API layer turns this
            # (or a name-hash fallback) into a portraits-folder URL.
            p["PLYR_PORTRAIT"] = match.get("PLYR_PORTRAIT")
            p["ovr"] = match.get("OverallRating")
            p["development_trait"] = match.get("TraitDevelopment")
            ptr = match.get("player_table_row")
            if ptr is not None and ptr in combine_data:
                p["combine_data"] = combine_data[ptr]
        grade_rec = (
            letter_grades.get(("flp", p.get("FirstName"), p.get("LastName"), p.get("position")))
            or letter_grades.get(("fl", p.get("FirstName"), p.get("LastName")))
        )
        if grade_rec:
            p["grades"] = grade_rec
            if not p.get("position"):
                p["position"] = grade_rec.get("position")
            p["height"] = grade_rec.get("height")
            p["weight"] = grade_rec.get("weight")
            p["age"] = grade_rec.get("age")

    _compute_combine_ranks(big_board["players"])

    loaded = {
        "year": str(year) if year is not None else None,
        "folder": str(folder),
        "folder_name": folder.name,
        "big_board": big_board,
        "players": players,
        "draft_picks": load_draft_picks(folder),
        "gm_info": load_gm_info(folder),
        "team_info": load_team_info(folder),
        "position_needs": load_position_needs(folder),
        "pick_values": load_draft_pick_value(folder),
        "max_per_position_group": load_max_per_position_group(folder),
        "colleges": colleges,
        "college_logo_map": college_logo_map,
        "nfl_logo_map": nfl_logo_map,
        "portrait_files": build_portrait_files(),
    }
    with _LOAD_ALL_CACHE_LOCK:
        _LOAD_ALL_CACHE.clear()
        _LOAD_ALL_CACHE[cache_key] = loaded
    return copy.deepcopy(loaded)
